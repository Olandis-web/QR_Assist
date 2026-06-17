# Importamos librerias para lecturas y decodificar los QR
import pyodbc
import flet as ft
import cv2
import threading
import time
import numpy as np
import openpyxl as xl
import os
from datetime import datetime
from pyzbar.pyzbar import decode
from database import conexion
from database.empleados import buscar_empleado
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment


class CamaraApp(ft.Container):
    
    def __init__(self, page, id_empleado_login):
        super().__init__(expand=True)
        self.bgcolor = ft.Colors.BLUE_GREY_900 
        self.alignment = ft.Alignment(0, 0)
        
        self.main_page = page
        self.id_empleado_login = id_empleado_login
        self.is_running = True
        self.capture = None

        self.registro_diario = []
        
        
        #Información mostrada en el panel lateral
        self.lbl_titulo = ft.Text("Información de Asistencia", color="white", size=22, weight="bold")
        self.lbl_fecha = ft.Text("Fecha: --/--/----", color="white70", size=16, weight="w500")
        self.lbl_hora = ft.Text("Hora: --:--:--", color="white70", size=16, weight="w500")
        
        self.lbl_resultado_titulo = ft.Text("Usuario Identificado:", color="white54", size=18)
        self.lbl_nombre = ft.Text("Esperando código QR...", color="yellow", size=24, weight="bold", text_align=ft.TextAlign.CENTER)
        self.lbl_mensaje = ft.Text("", size=18, weight="bold", text_align=ft.TextAlign.CENTER)
        
        # Contenedor que muestra los datos del empleado
        self.info_text = ft.Container(
            expand=True,
            border_radius=10,
            bgcolor=ft.Colors.BLUE_900,
            padding=30,
            content=ft.Column([
                self.lbl_titulo,
                ft.Divider(color="white24"),
                ft.Row([ft.Icon(ft.Icons.CALENDAR_TODAY, color="white70", size=18), self.lbl_fecha], spacing=10),
                ft.Row([ft.Icon(ft.Icons.ACCESS_TIME, color="white70", size=18), self.lbl_hora], spacing=10),
                ft.Divider(color="white24"),
                ft.Container(expand=True), 
                self.lbl_resultado_titulo,
                self.lbl_nombre,
                ft.Container(expand=True)
            ],
            alignment=ft.MainAxisAlignment.CENTER,
            horizontal_alignment=ft.CrossAxisAlignment.CENTER)
        )   
          
        self.content = ft.Row(
            controls=[self.info_text],
            expand=True
        
        )  
      
    def did_mount(self):
        """Esta función inicia el hilo de la cámara al cargar la vista"""
        if self.is_running and hasattr(self, 'threading') and self.threading.is_alive():
             return
        self.is_running = True
        self.threading = threading.Thread(target=self.update_frame_camara, daemon=True)
        self.threading.start()
      
    def infhora(self):
        """Esta función obtiene la fecha y hora actual"""
        inf = datetime.now()
        fecha = inf.strftime('%Y:%m:%d')
        hora = inf.strftime('%I:%M:%S %p')
        return hora, fecha
   
    def update_frame_camara(self):
        """Esta función captura el video, detecta QR y registra asistencias"""
        if self.capture is None or not self.capture.isOpened():
            self.capture = cv2.VideoCapture(0, cv2.CAP_DSHOW)   
            ret, frame = self.capture.read()
            if not ret:
                self.capture.release()
                self.capture = cv2.VideoCapture(0)
        
        while self.is_running:
            if self.capture is None or not self.capture.isOpened():
                time.sleep(1)
                continue
             
            ret, frame = self.capture.read()
            if not ret:
                time.sleep(0.01)
                continue
             
            # Dibuja el área de detección del QR
            cv2.putText(frame, 'Localizar el codigo QR', (160, 80), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 2)
            cv2.rectangle(frame, (170, 100), (470, 400), (0, 255, 0), 2)
    
            # Actualiza fecha y hora en pantalla
            hora, fecha = self.infhora()
            diasem = datetime.today().weekday()
             
            self.lbl_fecha.value = f"Fecha: {fecha.replace(':', '/')}"
            self.lbl_hora.value = f"Hora: {hora}"
            try:
                self.lbl_fecha.update()
                self.lbl_hora.update()
            except Exception:
                pass
            
            # Obtiene fecha y hora para clasificar la asistencia
            a, me, d = fecha[0:4], fecha[5:7], fecha[8:10]
            h = datetime.now().hour
            
            # Nombre del archivo Excel basado en la fecha actual
            nomar = f"{a}_{me}_{d}"
             
            # Lectura de códigos QR detectados
            for codes in decode(frame):
                try:
                    
                    info = codes.data.decode('utf-8')
                    codigo = int(info)

                    if codigo != self.id_empleado_login:
                        self.lbl_nombre.value = "QR no autorizado"
                        self.lbl_nombre.color = ft.Colors.RED
                        self.lbl_nombre.update()

                        continue
                    
                    pts = np.array([codes.polygon], np.int32)
                    xi, yi = codes.rect.left, codes.rect.top
                    pts = pts.reshape((-1, 1, 2))
                    cv2.polylines(frame, [pts], True, (255, 255, 0), 5)
                    
                    # Busca al empleado en la base de datos
                    empleado = buscar_empleado(codigo)
                    
                    # Muestra el nombre del empleado encontrado
                    if empleado:
                       nombre_completo = f"{empleado[0]} {empleado[1]}"

                       self.lbl_nombre.value = nombre_completo
                       self.lbl_nombre.color = ft.Colors.GREEN_ACCENT_400
                    else:
                        nombre_completo = f"ID{codigo}"
                       
                        self.lbl_nombre.value = "Empleado no encontrado"
                        self.lbl_nombre.color = ft.Colors.RED

                    self.lbl_nombre.update()
                 
                    # Registra de asistencia según el horario de Lunes a Viernes
                    if 4 >= diasem >= 0:

                        hora_actual = datetime.now().time()

                        if hora_actual.hour < 8:
                            estado = "A Tiempo"

                        elif hora_actual.hour == 8 and hora_actual.minute == 0:
                            estado = "A Tiempo"

                        else:
                            estado = "Tarde"

                        if codigo not in self.registro_diario:
                        
                            self.registro_diario.append(codigo)
                        
                        self._guardar_sql(codigo, estado)

                        self._guardar_excel(nomar, estado, nombre_completo)

                        cv2.putText(
                            frame,
                            f"{codigo} - {estado}",
                            (xi - 15, yi - 15),
                            cv2.FONT_HERSHEY_SIMPLEX,
                            1,
                            (255, 255, 0),
                            2
                        )

                                
                except Exception as e:
                    print(f"Error procesando QR: {e}")
            
            try:
                if self.is_running:
                    self.main_page.update()
            except Exception:
                pass    
            
            cv2.imshow("camara", frame)
            if cv2.waitKey(1) == 27 : #cierra con el teclado esc
                break
            
            time.sleep(0.03)
            
        self.is_running = False

        if self.capture is not None:
           self.capture.release()

        cv2.destroyAllWindows()

    def _guardar_excel(self, nombre_base, nombre_hoja, codigo):
        """Esta función guarda la asitencia en un archivo de Excel"""
        nombre_completo = f"{nombre_base}.xlsx"
        
         # Color utilizado para el encabezado
        color_encabezado = PatternFill(
                  start_color="1F4E78",
                  end_color="1F4E78",
                  fill_type="solid"
                )

         # Bordes aplicados a las celdas
        borde = Border(
                   left=Side(style="thin"),
                   right=Side(style="thin"),
                   top=Side(style="thin"),
                   bottom=Side(style="thin")
                )

        try:
            if os.path.exists(nombre_completo):
                wb = xl.load_workbook(nombre_completo)
            else:
                wb = xl.Workbook()
                if "Sheet" in wb.sheetnames:
                    wb.remove(wb["Sheet"])
            
            if nombre_hoja in wb.sheetnames:
                hoja = wb[nombre_hoja]
            else:
                hoja = wb.create_sheet(nombre_hoja)
                
                #Encabezados de la hoja
                hoja.append(["Empleado", "Hora Registro"])
                
               
                # Aplica formato al encabezado
                for celda in hoja[1]:
                    celda.fill = color_encabezado
                    celda.font = Font(bold=True, color="FFFFFF")
                    celda.alignment = Alignment(horizontal="center")
                    celda.border = borde
                    
            #Agrega registro de asistencia
            hoja.append([codigo, datetime.now().strftime('%I:%M:%S %p')])
           
            # Aplica bordes a la última fila agregada
            ultima_fila = hoja.max_row

            for celda in hoja[ultima_fila]:
                celda.border = borde
                celda.alignment = Alignment(horizontal="center")

            # Ajuste del tamaño de las columnas
            hoja.column_dimensions["A"].width = 35
            hoja.column_dimensions["B"].width = 18

            # Mantiene visible el encabezado al desplazarse
            hoja.freeze_panes = "A2"
            
            # Guarda los cambios realizados en el archivo
            wb.save(nombre_completo)
        except Exception as e:
            print(f"Error escribiendo en Excel: {e}")
    

    def _guardar_sql(self, id_empleado, estado):
        
        conn = conexion.conexion()
        cursor = conn.cursor()

        empleado = buscar_empleado(id_empleado)

        if not empleado:
            print(f"Empleado{id_empleado} no existe")
            conn.close()
            return

        cursor.execute("""SELECT COUNT(*) FROM Asistencia
                       WHERE ID_Empleado = ?
                       AND Fecha = CAST(GETDATE() AS DATE)""",(id_empleado,))
        
        existe = cursor.fetchone()[0]

        if existe > 0:
            conn.close()
            return
        
        cursor.execute("""INSERT INTO Asistencia
                       (
                        ID_Empleado,
                        Fecha,
                        Hora_entrada,
                        Estado
                       )

                       VALUES
                       (
                        ?,
                        CAST(GETDATE() AS DATE),
                        CAST(GETDATE() AS TIME),
                        ?
                       )
                       """, (id_empleado, estado))
        conn.commit()
        cursor.close()
        conn.close()
          
    def detener_camara(self):
        """Esta función detiene la cámara y libera los recursos utilizados"""
        self.is_running = False
        
        if hasattr(self, "capture") and self.capture:
          if self.capture.isOpened():
            self.capture.release()
          self.capture = None
          
        # Cierra todas las ventanas de OpenCV   
        cv2.destroyAllWindows()